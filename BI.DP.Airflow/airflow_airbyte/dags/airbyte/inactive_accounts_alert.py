import os
import logging
import pandas as pd
import snowflake.connector
from datetime import datetime
from airflow.models import Variable
from airflow.providers.snowflake.hooks.snowflake import SnowflakeHook
import base64
from io import BytesIO

logger = logging.getLogger(__name__)


def get_inactive_accounts_data():
    """
    Execute the Snowflake query to get accounts with no traffic for 90+ days
    Returns: pandas DataFrame with the query results
    """
    
    query = """
   WITH
AffTech AS
(SELECT
    TRACKER_LOGIN_ID,
    TRANSACTION_DATE,
    DATEDIFF(
        DAY,
        LAG(TRANSACTION_DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY TRANSACTION_DATE),
        TRANSACTION_DATE
    ) AS DIFFERENCE,
    CASE
        WHEN TRANSACTION_DATE = MAX(TRANSACTION_DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(TRANSACTION_DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Afftech' as Source
from RAW.AFFTECH.AFFILIATE_STATISTICS_REPORTS
--Where (DEPOSITS > 0 or NGR <> 0 or FTD_COUNT > 0 OR DEPOSIT_COUNT > 0 OR REGISTRATION > 0)
GROUP BY TRACKER_LOGIN_ID,TRANSACTION_DATE
),
Buffalo AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Buffalo Partners' as Source
from RAW.BUFFALO_PARTNERS.REV_SHARE_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--Where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT_AMOUNT > 0)
),
CellExpert AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'CellExpert' as Source
from RAW.CELLXPERT.DYNAMIC_VARIABLES_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--WHERE (DEPOSITS > 0 or PL <> 0)
),
Ego AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Ego' as Source
from RAW.EGO.BRAND_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--WHERE (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Income_Access AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Income Access' as Source
from RAW.INCOME_ACCESS.ACCOUNT_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Mexos AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Mexos' as Source
from RAW.MEXOS.STATISTICS_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
MyAffiliates AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'MyAffiliates' as Source
from RAW.MYAFFILIATES.CUSTOMER_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Netrefer AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Netrefer' as Source
from RAW.NETREFER.DYNAMIC_VARIABLES_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Q_Platform AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE(TRANSACTION_DATE) AS DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Q_Platform' as Source
from RAW.Q_PLATFORM.UTM_CODE_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Referon AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Referon' as Source
from RAW.REFERON.DYNAMIC_VARIABLES_REPORT
--where (DEPOSITS > 0 or NGR <> 0 or ftds_deposits > 0 or REG_COUNT > 0)
GROUP BY TRACKER_LOGIN_ID,DATE

),
Smartico AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Smartico' as Source
from RAW.SMARTICO.UTM_CODE_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),
Softswiss AS
(SELECT
    TRACKER_LOGIN_ID,
    DATE,
    DATEDIFF(
        DAY,
        LAG(DATE) OVER (PARTITION BY  TRACKER_LOGIN_ID ORDER BY DATE),
        DATE
    ) AS DIFFERENCE,
    CASE
        WHEN DATE = MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID)
        THEN DATEDIFF(
            DAY,
            MAX(DATE) OVER (PARTITION BY TRACKER_LOGIN_ID),
            CAST(DATEADD(DAY, -1, GETDATE()) AS date)
        )
        ELSE NULL
    END AS DIFF_FROM_YESTERDAY,
    'Softswiss' as Source
from RAW.SOFTSWISS.ACTIVITY_REPORT
GROUP BY TRACKER_LOGIN_ID,DATE
--where (DEPOSITS > 0 or NET_REVENUE <> 0 or FIRST_DEPOSIT > 0 or REGISTRATION_DATE is not null)
),

combined AS (
SELECT * FROM AffTech WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY >90
UNION ALL
SELECT * FROM Buffalo WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY >90
UNION ALL
SELECT * FROM CellExpert  WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY   >90
UNION ALL
SELECT * FROM Ego WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY  >90
UNION ALL
SELECT * FROM Income_Access WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY  >90
UNION ALL
SELECT * FROM Mexos WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY >90
UNION ALL
SELECT * FROM MyAffiliates WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY   >90
UNION ALL
SELECT * FROM Netrefer WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY   >90
UNION ALL
SELECT * FROM Q_Platform WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY  >90
UNION ALL
SELECT * FROM Referon WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY >90
UNION ALL
SELECT * FROM Smartico WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY  >90
UNION ALL
SELECT * FROM Softswiss WHERE DIFF_FROM_YESTERDAY IS NOT NULL AND DIFF_FROM_YESTERDAY  >90
)

    SELECT 
    Tracker_Login_ID,
    ID as BRT_ACCOUNT_ID,
    Username,
    Password,
    adve_name as Advertiser_name,
    PUBL_USERNAME as Publisher
    FROM combined C
    JOIN RAW.BRT.OPERATOR_ACCOUNTS Account
    ON Account.BR_TRACKER_LOGIN_ID = C.Tracker_Login_ID
    AND ACCOUNT.STATUS = 1
    JOIN RAW.BRC.TRACKER_LOGINS trk
    ON trk.tlog_id = account.br_tracker_login_id
    JOIN RAW.BRC.ADVERTISERS Adv
    ON trk.TLOG_FK_ADVERTISER = adv.ADVE_ID
    LEFT JOIN RAW.BRC.PUBLISHERS pub
    ON trk.TLOG_FK_PUBLISHER = pub.PUBL_ID
    ORDER BY Tracker_Login_ID
    """
    
    try:
        # Use Airflow's SnowflakeHook to connect using the configured connection
        snowflake_hook = SnowflakeHook(snowflake_conn_id='snowflake_conn')
        logger.info("Successfully connected to Snowflake database")
        
        # Execute query and fetch results
        result = snowflake_hook.get_records(query)
        
        # Get connection to access cursor description for column names
        conn = snowflake_hook.get_conn()
        cursor = conn.cursor()
        cursor.execute(query)
        column_names = [desc[0] for desc in cursor.description]
        cursor.close()
        
        logger.info(f"Query executed successfully, returned {len(result)} rows")
        
        # Convert to DataFrame
        df = pd.DataFrame(result, columns=column_names)
        
        return df
        
    except Exception as e:
        logger.error(f"Failed to execute Snowflake query: {e}")
        return None


def create_excel_report(df):
    """
    Create Excel file from DataFrame in memory and return as base64 encoded string
    Args:
        df: pandas DataFrame with the query results
    Returns:
        dict: Dictionary with 'filename', 'data' (base64 encoded), and 'size'
    """
    try:
        # Create timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inactive_accounts_report_{timestamp}.xlsx"
        
        logger.info(f"Creating Excel report in memory: {filename}")
        
        # Create Excel file in memory using BytesIO
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inactive Accounts', index=False)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets['Inactive Accounts']
            
            # Auto-size columns A and B to fit content without wrapping
            for column in ['A', 'B']:  # Tracker_Login_ID and BRT_ACCOUNT_ID
                max_length = 0
                for cell in worksheet[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width
            
            # Set fixed narrower widths for remaining columns
            worksheet.column_dimensions['C'].width = 18  # Username
            worksheet.column_dimensions['D'].width = 15  # Password
            worksheet.column_dimensions['E'].width = 22  # Advertiser_name
            worksheet.column_dimensions['F'].width = 20  # Publisher
            
            # Set page setup for better Slack preview
            worksheet.page_setup.orientation = 'landscape'
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 0  # Allow multiple pages vertically if needed
            
            # Add borders and styling for better readability
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            # Style header row with smaller font
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=9)  # Smaller header font
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add borders to all cells with smaller font for data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            data_font = Font(size=8)  # Smaller data font
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        # Get the binary data
        excel_bytes = output.getvalue()
        file_size = len(excel_bytes)
        
        logger.info(f"Excel file created in memory, size: {file_size} bytes ({file_size / 1024:.2f} KB)")
        
        # Encode as base64 for XCom transport
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        encoded_size = len(excel_base64)
        
        logger.info(f"Base64 encoded size: {encoded_size} bytes ({encoded_size / 1024:.2f} KB)")
        logger.info("Excel report created successfully in memory")
        
        # Return dictionary with filename and data
        return {
            'filename': filename,
            'data': excel_base64,
            'size': file_size
        }
        
    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def execute_inactive_accounts_report():
    """
    Main function to execute the inactive accounts report process
    Returns:
        dict: Dictionary with 'filename', 'data', and 'size', or None if failed
    """
    try:
        logger.info("Starting inactive accounts report generation")
        
        # Get data from Snowflake
        df = get_inactive_accounts_data()
        
        if df is None or df.empty:
            logger.warning("No inactive accounts found or query failed")
            return None
        
        logger.info(f"Found {len(df)} inactive accounts")
        
        # Create Excel report (returns dict with filename and base64 data)
        excel_data = create_excel_report(df)
        
        if excel_data:
            logger.info("Inactive accounts report generated successfully")
            logger.info(f"Report will be sent as: {excel_data['filename']}")
            return excel_data
        else:
            logger.error("Failed to create Excel report")
            return None
            
    except Exception as e:
        logger.error(f"Error in execute_inactive_accounts_report: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
